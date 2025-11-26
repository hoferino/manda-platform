"""
Excel-specific parser with formula preservation.
Story: E3.2 - Integrate Docling for Document Parsing (AC: #2)

This module handles Excel files (.xlsx, .xls) with special handling for:
- Multi-sheet workbooks (each sheet becomes separate chunks)
- Formula preservation as text (e.g., "=SUM(A1:A10)")
- Table extraction with row/column structure
- Cell reference tracking in metadata
- Named ranges and cell references
"""

import re
from pathlib import Path
from typing import Any, Optional

import structlog
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.config import Settings, get_settings
from src.parsers import (
    ChunkData,
    CorruptFileError,
    FormulaData,
    ParseError,
    ParseResult,
    TableData,
)
from src.parsers.chunker import Chunker, get_chunker

logger = structlog.get_logger(__name__)


class ExcelParser:
    """
    Excel parser with formula preservation using openpyxl.

    This parser is used instead of Docling for Excel files because:
    1. Docling doesn't preserve Excel formulas
    2. We need detailed cell reference metadata
    3. We need to identify and extract tables separately
    """

    def __init__(
        self,
        config: Optional[Settings] = None,
        chunker: Optional[Chunker] = None,
    ):
        """
        Initialize the Excel parser.

        Args:
            config: Application settings
            chunker: Chunker instance for text splitting
        """
        self.config = config or get_settings()
        self.chunker = chunker or get_chunker()

    def supports(self, file_type: str) -> bool:
        """Check if this parser supports the given file type."""
        file_type_lower = file_type.lower()
        return any(
            ext in file_type_lower
            for ext in [".xlsx", ".xls", "spreadsheet", "excel"]
        )

    async def parse(
        self,
        file_path: Path,
        file_type: str = "xlsx",
    ) -> ParseResult:
        """
        Parse an Excel file and return structured content.

        Args:
            file_path: Path to the Excel file
            file_type: MIME type or extension

        Returns:
            ParseResult with chunks (per sheet), tables, and formulas
        """
        chunks: list[ChunkData] = []
        tables: list[TableData] = []
        formulas: list[FormulaData] = []
        errors: list[str] = []
        warnings: list[str] = []
        total_sheets = 0  # Initialize before try block to avoid scope issues

        chunk_index = 0

        try:
            # Load workbook with data_only=False to preserve formulas
            wb = load_workbook(
                filename=str(file_path),
                data_only=False,
                read_only=False,
            )

            # Also load with data_only=True to get computed values
            try:
                wb_values = load_workbook(
                    filename=str(file_path),
                    data_only=True,
                    read_only=True,
                )
            except Exception:
                wb_values = None
                warnings.append("Could not load computed values; formula results may be unavailable")

            total_sheets = len(wb.sheetnames)

            logger.info(
                "Processing Excel workbook",
                file=str(file_path),
                sheets=total_sheets,
            )

            # Process each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                value_sheet = wb_values[sheet_name] if wb_values else None

                try:
                    # Extract sheet content as structured data
                    sheet_content, sheet_tables, sheet_formulas = self._process_sheet(
                        sheet,
                        value_sheet,
                        sheet_name,
                        file_path.name,
                    )

                    # Create chunk(s) for sheet content
                    if sheet_content.strip():
                        sheet_chunks = self.chunker.chunk_text(
                            sheet_content,
                            chunk_type="text",
                            base_metadata={
                                "source_file": file_path.name,
                                "sheet_name": sheet_name,
                            },
                            sheet_name=sheet_name,
                        )
                        for chunk in sheet_chunks:
                            chunk.chunk_index = chunk_index
                            chunks.append(chunk)
                            chunk_index += 1

                    # Add tables (also create chunks for them)
                    for table in sheet_tables:
                        tables.append(table)

                        # Create chunk for table
                        table_chunks = self.chunker.chunk_table(
                            table.content,
                            chunk_index=chunk_index,
                            base_metadata={
                                "source_file": file_path.name,
                                "sheet_name": sheet_name,
                                "is_table": True,
                            },
                            sheet_name=sheet_name,
                        )
                        for chunk in table_chunks:
                            chunks.append(chunk)
                            chunk_index += 1

                    # Add formulas
                    formulas.extend(sheet_formulas)

                except Exception as e:
                    errors.append(f"Error processing sheet '{sheet_name}': {str(e)}")
                    logger.warning(
                        "Sheet processing error",
                        sheet=sheet_name,
                        error=str(e),
                    )

            # Create formula chunks for significant formulas
            if formulas:
                formula_content = self._format_formulas_as_text(formulas)
                if formula_content:
                    formula_chunk = ChunkData(
                        content=formula_content,
                        chunk_type="formula",
                        chunk_index=chunk_index,
                        metadata={
                            "source_file": file_path.name,
                            "formula_count": len(formulas),
                        },
                    )
                    chunks.append(formula_chunk)

            wb.close()
            if wb_values:
                wb_values.close()

        except Exception as e:
            if isinstance(e, ParseError):
                raise

            logger.error(
                "Excel parsing failed",
                file=str(file_path),
                error=str(e),
                exc_info=True,
            )
            raise CorruptFileError(
                f"Failed to parse Excel file: {str(e)}",
                file_path=file_path,
                details={"original_error": str(e)},
            )

        return ParseResult(
            chunks=chunks,
            tables=tables,
            formulas=formulas,
            metadata={
                "source": str(file_path),
                "total_sheets": total_sheets,
            },
            total_sheets=total_sheets,
            errors=errors,
            warnings=warnings,
        )

    def _process_sheet(
        self,
        sheet: Worksheet,
        value_sheet: Optional[Worksheet],
        sheet_name: str,
        source_file: str,
    ) -> tuple[str, list[TableData], list[FormulaData]]:
        """
        Process a single Excel sheet.

        Returns:
            Tuple of (text content, tables, formulas)
        """
        text_parts: list[str] = []
        tables: list[TableData] = []
        formulas: list[FormulaData] = []

        # Get dimensions
        min_row = sheet.min_row or 1
        max_row = sheet.max_row or 1
        min_col = sheet.min_column or 1
        max_col = sheet.max_column or 1

        if max_row - min_row > 10000:
            logger.warning(
                "Large sheet detected, may be slow",
                sheet=sheet_name,
                rows=max_row - min_row,
            )

        # Detect tables (contiguous regions with headers)
        detected_tables = self._detect_tables(sheet, min_row, max_row, min_col, max_col)

        for table_region in detected_tables:
            table_data = self._extract_table(
                sheet,
                table_region,
                sheet_name,
                source_file,
            )
            if table_data:
                tables.append(table_data)

        # Extract all formulas
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    # Get computed value from value_sheet if available
                    result_value = None
                    if value_sheet:
                        try:
                            value_cell = value_sheet.cell(row=cell.row, column=cell.column)
                            if value_cell.value is not None:
                                result_value = str(value_cell.value)
                        except Exception:
                            pass

                    formula_data = FormulaData(
                        formula=cell.value,
                        cell_reference=f"{get_column_letter(cell.column)}{cell.row}",
                        sheet_name=sheet_name,
                        result_value=result_value,
                        references=self._extract_cell_references(cell.value),
                    )
                    formulas.append(formula_data)

        # Build text representation of sheet (non-table content)
        text_content = self._sheet_to_text(sheet, min_row, max_row, min_col, max_col)
        if text_content:
            text_parts.append(f"## Sheet: {sheet_name}\n\n{text_content}")

        return "\n\n".join(text_parts), tables, formulas

    def _detect_tables(
        self,
        sheet: Worksheet,
        min_row: int,
        max_row: int,
        min_col: int,
        max_col: int,
    ) -> list[dict]:
        """
        Detect table regions in the sheet.

        A table is identified by:
        - A row with text values (headers)
        - Followed by rows with consistent data patterns
        """
        tables: list[dict] = []

        # Simple heuristic: look for rows that look like headers
        # (text in multiple consecutive cells)
        for row_num in range(min_row, min(max_row, min_row + 100)):  # Check first 100 rows
            header_cells = []
            for col_num in range(min_col, max_col + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                if cell.value is not None and isinstance(cell.value, str):
                    header_cells.append((col_num, cell.value))
                elif cell.value is not None:
                    header_cells.append((col_num, str(cell.value)))

            # If we have 3+ consecutive header-like cells, might be a table
            if len(header_cells) >= 3:
                # Check if next rows have data
                data_rows = 0
                for check_row in range(row_num + 1, min(max_row + 1, row_num + 50)):
                    has_data = False
                    for col_num, _ in header_cells:
                        check_cell = sheet.cell(row=check_row, column=col_num)
                        if check_cell.value is not None:
                            has_data = True
                            break
                    if has_data:
                        data_rows += 1
                    else:
                        break

                if data_rows >= 2:  # At least 2 data rows
                    tables.append({
                        "start_row": row_num,
                        "end_row": row_num + data_rows,
                        "start_col": min(c[0] for c in header_cells),
                        "end_col": max(c[0] for c in header_cells),
                    })

        return tables

    def _extract_table(
        self,
        sheet: Worksheet,
        region: dict,
        sheet_name: str,
        source_file: str,
    ) -> Optional[TableData]:
        """Extract a table from the specified region."""
        try:
            rows_data: list[list[str]] = []
            headers: list[str] = []

            start_row = region["start_row"]
            end_row = region["end_row"]
            start_col = region["start_col"]
            end_col = region["end_col"]

            # Extract headers (first row)
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=start_row, column=col)
                headers.append(str(cell.value) if cell.value else "")

            # Extract data rows
            for row_num in range(start_row + 1, end_row + 1):
                row_data: list[str] = []
                for col in range(start_col, end_col + 1):
                    cell = sheet.cell(row=row_num, column=col)
                    row_data.append(str(cell.value) if cell.value else "")
                rows_data.append(row_data)

            # Build markdown table
            markdown = self._build_markdown_table(headers, rows_data)

            return TableData(
                content=markdown,
                rows=len(rows_data) + 1,  # Include header
                cols=len(headers),
                headers=headers,
                sheet_name=sheet_name,
                data=rows_data,
            )

        except Exception as e:
            logger.debug("Error extracting table", error=str(e), region=region)
            return None

    def _build_markdown_table(
        self,
        headers: list[str],
        rows: list[list[str]],
    ) -> str:
        """Build a markdown table from headers and rows."""
        if not headers:
            return ""

        lines: list[str] = []

        # Header row
        header_line = "| " + " | ".join(h or " " for h in headers) + " |"
        lines.append(header_line)

        # Separator
        separator = "| " + " | ".join("---" for _ in headers) + " |"
        lines.append(separator)

        # Data rows
        for row in rows:
            # Pad row if needed
            padded_row = row + [""] * (len(headers) - len(row))
            row_line = "| " + " | ".join(cell or " " for cell in padded_row[:len(headers)]) + " |"
            lines.append(row_line)

        return "\n".join(lines)

    def _sheet_to_text(
        self,
        sheet: Worksheet,
        min_row: int,
        max_row: int,
        min_col: int,
        max_col: int,
    ) -> str:
        """Convert sheet content to plain text representation."""
        lines: list[str] = []

        for row_num in range(min_row, min(max_row + 1, min_row + 1000)):  # Limit rows
            row_values: list[str] = []
            has_content = False

            for col_num in range(min_col, min(max_col + 1, min_col + 50)):  # Limit columns
                cell = sheet.cell(row=row_num, column=col_num)
                if cell.value is not None:
                    has_content = True
                    # Show formula as text if present
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        row_values.append(f"[Formula: {cell.value}]")
                    else:
                        row_values.append(str(cell.value))
                else:
                    row_values.append("")

            if has_content:
                # Only include non-empty values
                content = "\t".join(v for v in row_values if v)
                if content.strip():
                    lines.append(content)

        return "\n".join(lines)

    def _extract_cell_references(self, formula: str) -> list[str]:
        """
        Extract cell references from a formula.

        Examples:
            "=SUM(A1:A10)" -> ["A1", "A10"]
            "=B5+C5*D5" -> ["B5", "C5", "D5"]
        """
        # Pattern to match cell references like A1, B23, AA100, Sheet1!A1
        pattern = r"(?:\'?[\w\s]+\'?!)?\$?[A-Z]{1,3}\$?\d+"
        matches = re.findall(pattern, formula.upper())
        return list(set(matches))  # Remove duplicates

    def _format_formulas_as_text(self, formulas: list[FormulaData]) -> str:
        """Format formulas as a readable text summary."""
        if not formulas:
            return ""

        lines = ["## Formula Summary\n"]

        # Group by sheet
        by_sheet: dict[str, list[FormulaData]] = {}
        for f in formulas:
            if f.sheet_name not in by_sheet:
                by_sheet[f.sheet_name] = []
            by_sheet[f.sheet_name].append(f)

        for sheet_name, sheet_formulas in by_sheet.items():
            lines.append(f"\n### Sheet: {sheet_name}\n")

            # Only include significant formulas (not simple references)
            significant = [
                f for f in sheet_formulas
                if any(op in f.formula.upper() for op in ["SUM", "AVERAGE", "IF", "VLOOKUP", "INDEX", "MATCH", "+", "-", "*", "/"])
            ]

            for f in significant[:50]:  # Limit to 50 per sheet
                result = f" = {f.result_value}" if f.result_value else ""
                lines.append(f"- **{f.cell_reference}**: `{f.formula}`{result}")

        return "\n".join(lines)


def create_excel_parser(
    config: Optional[Settings] = None,
    chunker: Optional[Chunker] = None,
) -> ExcelParser:
    """Create an ExcelParser instance."""
    return ExcelParser(config=config, chunker=chunker)


__all__ = ["ExcelParser", "create_excel_parser"]
