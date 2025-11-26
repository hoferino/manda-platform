"""
Unit tests for Docling parser wrapper.
Story: E3.2 - Integrate Docling for Document Parsing (AC: #1, #5, #6)

Note: These tests require the docling library to be installed.
They will be skipped if docling is not available.
"""

import pytest
from pathlib import Path
from unittest.mock import MagicMock, patch, AsyncMock
import tempfile
import os

# Check if docling is available
try:
    import docling
    DOCLING_AVAILABLE = True
except ImportError:
    DOCLING_AVAILABLE = False

pytestmark = pytest.mark.skipif(
    not DOCLING_AVAILABLE,
    reason="docling library not installed - tests require docling>=2.15.0"
)


class TestDoclingParserBasic:
    """Basic tests for DoclingParser class."""

    def test_docling_parser_initialization(self, mock_parser_settings: MagicMock) -> None:
        """Test that DoclingParser initializes correctly."""
        with patch("src.parsers.docling_parser.DocumentConverter"):
            from src.parsers.docling_parser import DoclingParser

            parser = DoclingParser()

            assert parser.config is not None
            assert parser.chunker is not None

    def test_supports_all_types(self, mock_parser_settings: MagicMock) -> None:
        """Test that DoclingParser supports all expected types."""
        with patch("src.parsers.docling_parser.DocumentConverter"):
            from src.parsers.docling_parser import DoclingParser

            parser = DoclingParser()

            # PDF
            assert parser.supports("application/pdf") is True
            assert parser.supports(".pdf") is True

            # Excel
            assert parser.supports("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet") is True
            assert parser.supports(".xlsx") is True

            # Word
            assert parser.supports("application/vnd.openxmlformats-officedocument.wordprocessingml.document") is True
            assert parser.supports(".docx") is True

            # Images
            assert parser.supports("image/png") is True
            assert parser.supports(".jpg") is True

            # Unsupported
            assert parser.supports(".txt") is False
            assert parser.supports(".csv") is False


class TestDoclingParserValidation:
    """Tests for input validation."""

    @pytest.mark.asyncio
    async def test_parse_unsupported_type(self, mock_parser_settings: MagicMock) -> None:
        """Test that unsupported file types raise error."""
        with patch("src.parsers.docling_parser.DocumentConverter"):
            from src.parsers.docling_parser import DoclingParser
            from src.parsers import UnsupportedFileTypeError

            parser = DoclingParser()

            with tempfile.NamedTemporaryFile(suffix=".txt", delete=False) as f:
                f.write(b"text content")
                temp_path = Path(f.name)

            try:
                with pytest.raises(UnsupportedFileTypeError):
                    await parser.parse(temp_path, "text/plain")
            finally:
                os.unlink(temp_path)

    @pytest.mark.asyncio
    async def test_parse_nonexistent_file(self, mock_parser_settings: MagicMock) -> None:
        """Test that non-existent files raise error."""
        with patch("src.parsers.docling_parser.DocumentConverter"):
            from src.parsers.docling_parser import DoclingParser
            from src.parsers import ParseError

            parser = DoclingParser()

            with pytest.raises(ParseError):
                await parser.parse(Path("/nonexistent/file.pdf"), "application/pdf")

    @pytest.mark.asyncio
    async def test_parse_file_too_large(self, mock_parser_settings: MagicMock) -> None:
        """Test that oversized files raise error."""
        # Override max size to 0 to trigger error
        mock_parser_settings.parser_max_file_size_mb = 0

        with patch("src.parsers.docling_parser.DocumentConverter"):
            from src.parsers.docling_parser import DoclingParser
            from src.parsers import FileTooLargeError

            parser = DoclingParser()

            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
                f.write(b"X" * 1024)  # 1KB file
                temp_path = Path(f.name)

            try:
                with pytest.raises(FileTooLargeError):
                    await parser.parse(temp_path, "application/pdf")
            finally:
                os.unlink(temp_path)


class TestDoclingParserPDF:
    """Tests for PDF parsing through DoclingParser."""

    @pytest.mark.asyncio
    async def test_parse_pdf_delegates_to_processor(
        self,
        mock_parser_settings: MagicMock,
    ) -> None:
        """Test that PDF parsing produces expected result structure."""
        # Mock the DocumentConverter
        mock_doc = MagicMock()
        mock_doc.pages = None
        mock_doc.export_to_markdown.return_value = "PDF content here"
        mock_doc.tables = []

        mock_result = MagicMock()
        mock_result.document = mock_doc

        mock_converter = MagicMock()
        mock_converter.convert.return_value = mock_result

        with patch("src.parsers.docling_parser.DocumentConverter", return_value=mock_converter):
            from src.parsers.docling_parser import DoclingParser

            parser = DoclingParser()

            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
                f.write(b"%PDF-1.4 content")
                temp_path = Path(f.name)

            try:
                result = await parser.parse(temp_path, "application/pdf")

                assert len(result.chunks) > 0
                assert result.metadata.get("file_category") == "pdf"
                assert result.formulas == []

            finally:
                os.unlink(temp_path)


class TestDoclingParserExcel:
    """Tests for Excel parsing through DoclingParser."""

    @pytest.mark.asyncio
    async def test_parse_excel_uses_excel_parser(
        self,
        mock_parser_settings: MagicMock,
    ) -> None:
        """Test that Excel files are delegated to ExcelParser."""
        from openpyxl import Workbook

        # Create real Excel file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = Path(f.name)

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test"
        ws["B1"] = "=A1"
        wb.save(temp_path)

        try:
            with patch("src.parsers.docling_parser.DocumentConverter"):
                from src.parsers.docling_parser import DoclingParser

                parser = DoclingParser()
                result = await parser.parse(temp_path, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                # Excel should have formulas
                assert len(result.chunks) > 0
                assert len(result.formulas) > 0  # Should find the =A1 formula
                assert result.formulas[0].formula == "=A1"

        finally:
            os.unlink(temp_path)


class TestDoclingParserWord:
    """Tests for Word parsing through DoclingParser."""

    @pytest.mark.asyncio
    async def test_parse_word_document(
        self,
        mock_parser_settings: MagicMock,
    ) -> None:
        """Test parsing Word documents."""
        mock_doc = MagicMock()
        mock_doc.export_to_markdown.return_value = """
        # Document Title

        This is a Word document with some content.

        ## Section 1

        Important information here.
        """
        mock_doc.tables = []

        mock_result = MagicMock()
        mock_result.document = mock_doc

        mock_converter = MagicMock()
        mock_converter.convert.return_value = mock_result

        with patch("src.parsers.docling_parser.DocumentConverter", return_value=mock_converter):
            from src.parsers.docling_parser import DoclingParser

            parser = DoclingParser()

            with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as f:
                f.write(b"PK")  # DOCX files start with PK (zip signature)
                temp_path = Path(f.name)

            try:
                result = await parser.parse(temp_path, "application/vnd.openxmlformats-officedocument.wordprocessingml.document")

                assert len(result.chunks) > 0
                assert result.metadata.get("file_category") == "word"
                combined = " ".join(c.content for c in result.chunks)
                assert "Document" in combined or "information" in combined

            finally:
                os.unlink(temp_path)


class TestDoclingParserImage:
    """Tests for image OCR through DoclingParser."""

    @pytest.mark.asyncio
    async def test_parse_image_with_ocr(
        self,
        mock_parser_settings: MagicMock,
    ) -> None:
        """Test OCR processing of images."""
        mock_doc = MagicMock()
        mock_doc.export_to_markdown.return_value = "OCR extracted text from image"

        mock_result = MagicMock()
        mock_result.document = mock_doc

        mock_converter = MagicMock()
        mock_converter.convert.return_value = mock_result

        with patch("src.parsers.docling_parser.DocumentConverter", return_value=mock_converter):
            from src.parsers.docling_parser import DoclingParser

            parser = DoclingParser()

            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
                # Write minimal PNG header
                f.write(b'\x89PNG\r\n\x1a\n')
                temp_path = Path(f.name)

            try:
                result = await parser.parse(temp_path, "image/png")

                assert len(result.chunks) > 0
                assert result.metadata.get("ocr_processed") is True
                # Chunks should be marked as image-sourced
                assert any(c.chunk_type == "image" for c in result.chunks)

            finally:
                os.unlink(temp_path)


class TestDoclingParserMetadata:
    """Tests for metadata handling."""

    @pytest.mark.asyncio
    async def test_parse_records_timing(
        self,
        mock_parser_settings: MagicMock,
    ) -> None:
        """Test that parse time is recorded."""
        mock_doc = MagicMock()
        mock_doc.pages = None
        mock_doc.export_to_markdown.return_value = "Content"

        mock_result = MagicMock()
        mock_result.document = mock_doc

        mock_converter = MagicMock()
        mock_converter.convert.return_value = mock_result

        with patch("src.parsers.docling_parser.DocumentConverter", return_value=mock_converter):
            from src.parsers.docling_parser import DoclingParser

            parser = DoclingParser()

            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
                f.write(b"%PDF content")
                temp_path = Path(f.name)

            try:
                result = await parser.parse(temp_path, "application/pdf")

                assert result.parse_time_ms is not None
                assert result.parse_time_ms >= 0
                assert result.metadata.get("file_type") == "application/pdf"
                assert result.metadata.get("file_size_bytes") > 0

            finally:
                os.unlink(temp_path)


class TestDoclingParserFactory:
    """Tests for factory function."""

    def test_create_docling_parser(self, mock_parser_settings: MagicMock) -> None:
        """Test create_docling_parser factory."""
        with patch("src.parsers.docling_parser.DocumentConverter"):
            from src.parsers.docling_parser import create_docling_parser

            parser = create_docling_parser()

            assert parser is not None
            assert hasattr(parser, "parse")
            assert hasattr(parser, "supports")

    def test_create_docling_parser_with_custom_config(
        self,
        mock_parser_settings: MagicMock,
    ) -> None:
        """Test factory with custom config."""
        with patch("src.parsers.docling_parser.DocumentConverter"):
            from src.parsers.docling_parser import create_docling_parser
            from src.parsers.chunker import Chunker

            custom_chunker = Chunker(max_tokens=500)

            parser = create_docling_parser(chunker=custom_chunker)

            assert parser.chunker.max_tokens == 500
