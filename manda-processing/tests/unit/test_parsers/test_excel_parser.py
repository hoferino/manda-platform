"""
Unit tests for Excel parser with formula preservation.
Story: E3.2 - Integrate Docling for Document Parsing (AC: #2, #6)
"""

import pytest
from pathlib import Path
from unittest.mock import MagicMock, patch, AsyncMock
from io import BytesIO
import tempfile
import os


class TestExcelParserBasic:
    """Basic tests for ExcelParser class."""

    def test_excel_parser_initialization(self, mock_parser_settings: MagicMock) -> None:
        """Test that ExcelParser initializes correctly."""
        from src.parsers.excel_parser import ExcelParser

        parser = ExcelParser()

        assert parser.config is not None
        assert parser.chunker is not None

    def test_supports_excel_types(self, mock_parser_settings: MagicMock) -> None:
        """Test that parser supports Excel file types."""
        from src.parsers.excel_parser import ExcelParser

        parser = ExcelParser()

        assert parser.supports(".xlsx") is True
        assert parser.supports(".xls") is True
        assert parser.supports("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet") is True
        assert parser.supports("spreadsheet") is True
        assert parser.supports(".pdf") is False
        assert parser.supports(".docx") is False


class TestExcelFormulaExtraction:
    """Tests for Excel formula extraction."""

    def test_extract_cell_references_sum(self, mock_parser_settings: MagicMock) -> None:
        """Test extracting cell references from SUM formula."""
        from src.parsers.excel_parser import ExcelParser

        parser = ExcelParser()

        refs = parser._extract_cell_references("=SUM(A1:A10)")
        assert "A1" in refs
        assert "A10" in refs

    def test_extract_cell_references_complex(self, mock_parser_settings: MagicMock) -> None:
        """Test extracting cell references from complex formula."""
        from src.parsers.excel_parser import ExcelParser

        parser = ExcelParser()

        refs = parser._extract_cell_references("=B5*C5+D5/E5")
        assert "B5" in refs
        assert "C5" in refs
        assert "D5" in refs
        assert "E5" in refs

    def test_extract_cell_references_vlookup(self, mock_parser_settings: MagicMock) -> None:
        """Test extracting cell references from VLOOKUP."""
        from src.parsers.excel_parser import ExcelParser

        parser = ExcelParser()

        refs = parser._extract_cell_references("=VLOOKUP(A1,Sheet2!A:B,2,FALSE)")
        assert "A1" in refs
        # Sheet references should be captured too
        assert any("A" in ref or "B" in ref for ref in refs)

    def test_extract_cell_references_if_formula(self, mock_parser_settings: MagicMock) -> None:
        """Test extracting cell references from IF formula."""
        from src.parsers.excel_parser import ExcelParser

        parser = ExcelParser()

        refs = parser._extract_cell_references('=IF(A1>100,"High","Low")')
        assert "A1" in refs


class TestExcelTableExtraction:
    """Tests for Excel table detection and extraction."""

    def test_build_markdown_table(self, mock_parser_settings: MagicMock) -> None:
        """Test markdown table building."""
        from src.parsers.excel_parser import ExcelParser

        parser = ExcelParser()

        headers = ["Name", "Value", "Change"]
        rows = [
            ["Revenue", "150M", "+25%"],
            ["EBITDA", "27M", "+50%"],
        ]

        markdown = parser._build_markdown_table(headers, rows)

        assert "| Name | Value | Change |" in markdown
        assert "| --- | --- | --- |" in markdown
        assert "| Revenue | 150M | +25% |" in markdown
        assert "| EBITDA | 27M | +50% |" in markdown

    def test_build_markdown_table_empty_headers(self, mock_parser_settings: MagicMock) -> None:
        """Test handling empty headers."""
        from src.parsers.excel_parser import ExcelParser

        parser = ExcelParser()

        result = parser._build_markdown_table([], [])
        assert result == ""

    def test_build_markdown_table_mismatched_rows(self, mock_parser_settings: MagicMock) -> None:
        """Test handling rows with different lengths than headers."""
        from src.parsers.excel_parser import ExcelParser

        parser = ExcelParser()

        headers = ["A", "B", "C"]
        rows = [
            ["1", "2"],  # Missing one column
            ["4", "5", "6", "7"],  # Extra column
        ]

        markdown = parser._build_markdown_table(headers, rows)

        # Should pad/truncate to match headers
        assert "| A | B | C |" in markdown
        lines = markdown.split("\n")
        # All data lines should have correct column count
        for line in lines[2:]:  # Skip header and separator
            assert line.count("|") == 4  # 3 columns + outer pipes


class TestExcelFormulaSummary:
    """Tests for formula summary generation."""

    def test_format_formulas_as_text(
        self,
        mock_parser_settings: MagicMock,
        sample_formulas: list[tuple[str, str, str]],
    ) -> None:
        """Test formatting formulas as readable text."""
        from src.parsers.excel_parser import ExcelParser
        from src.parsers import FormulaData

        parser = ExcelParser()

        formulas = [
            FormulaData(
                formula=f[0],
                cell_reference=f[1],
                sheet_name=f[2],
            )
            for f in sample_formulas
        ]

        text = parser._format_formulas_as_text(formulas)

        assert "## Formula Summary" in text
        assert "=SUM(A1:A10)" in text
        assert "Sheet1" in text or "Revenue" in text

    def test_format_formulas_empty(self, mock_parser_settings: MagicMock) -> None:
        """Test handling empty formula list."""
        from src.parsers.excel_parser import ExcelParser

        parser = ExcelParser()

        result = parser._format_formulas_as_text([])
        assert result == ""


class TestExcelParserWithRealFile:
    """Integration tests using real Excel files (created in test)."""

    @pytest.fixture
    def simple_excel_file(self, tmp_path: Path) -> Path:
        """Create a simple Excel file for testing."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Add some data
        ws["A1"] = "Item"
        ws["B1"] = "Value"
        ws["A2"] = "Revenue"
        ws["B2"] = 1000000
        ws["A3"] = "Costs"
        ws["B3"] = 600000
        ws["A4"] = "Profit"
        ws["B4"] = "=B2-B3"  # Formula

        file_path = tmp_path / "test.xlsx"
        wb.save(file_path)
        return file_path

    @pytest.fixture
    def multi_sheet_excel(self, tmp_path: Path) -> Path:
        """Create a multi-sheet Excel file."""
        from openpyxl import Workbook

        wb = Workbook()

        # Sheet 1: Summary
        ws1 = wb.active
        ws1.title = "Summary"
        ws1["A1"] = "Metric"
        ws1["B1"] = "Value"
        ws1["A2"] = "Total Revenue"
        ws1["B2"] = "=Revenue!B10"

        # Sheet 2: Revenue
        ws2 = wb.create_sheet("Revenue")
        ws2["A1"] = "Month"
        ws2["B1"] = "Amount"
        for i in range(2, 11):
            ws2[f"A{i}"] = f"Month {i-1}"
            ws2[f"B{i}"] = 100000 * (i - 1)
        ws2["B10"] = "=SUM(B2:B9)"

        # Sheet 3: Expenses
        ws3 = wb.create_sheet("Expenses")
        ws3["A1"] = "Category"
        ws3["B1"] = "Amount"
        ws3["A2"] = "Salaries"
        ws3["B2"] = 50000

        file_path = tmp_path / "multi_sheet.xlsx"
        wb.save(file_path)
        return file_path

    @pytest.mark.asyncio
    async def test_parse_simple_excel(
        self,
        mock_parser_settings: MagicMock,
        simple_excel_file: Path,
    ) -> None:
        """Test parsing a simple Excel file."""
        from src.parsers.excel_parser import ExcelParser

        parser = ExcelParser()
        result = await parser.parse(simple_excel_file)

        assert len(result.chunks) > 0
        assert len(result.formulas) == 1
        assert result.formulas[0].formula == "=B2-B3"
        assert result.formulas[0].cell_reference == "B4"

    @pytest.mark.asyncio
    async def test_parse_multi_sheet(
        self,
        mock_parser_settings: MagicMock,
        multi_sheet_excel: Path,
    ) -> None:
        """Test parsing a multi-sheet Excel file."""
        from src.parsers.excel_parser import ExcelParser

        parser = ExcelParser()
        result = await parser.parse(multi_sheet_excel)

        assert len(result.chunks) > 0
        assert result.total_sheets == 3

        # Check we got formulas from multiple sheets
        sheets_with_formulas = set(f.sheet_name for f in result.formulas)
        assert len(sheets_with_formulas) >= 1

        # Check sheet names in chunks
        chunk_sheets = set(
            c.sheet_name for c in result.chunks if c.sheet_name
        )
        assert "Summary" in chunk_sheets or "Revenue" in chunk_sheets

    @pytest.mark.asyncio
    async def test_parse_extracts_tables(
        self,
        mock_parser_settings: MagicMock,
        tmp_path: Path,
    ) -> None:
        """Test that tables are detected and extracted."""
        from openpyxl import Workbook
        from src.parsers.excel_parser import ExcelParser

        # Create a file with a proper table structure
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales"

        # Add headers and data (table-like structure)
        ws["A1"] = "Product"
        ws["B1"] = "Q1"
        ws["C1"] = "Q2"
        ws["D1"] = "Q3"
        for i in range(2, 10):
            ws[f"A{i}"] = f"Product {i-1}"
            ws[f"B{i}"] = 1000 * i
            ws[f"C{i}"] = 1200 * i
            ws[f"D{i}"] = 1500 * i

        file_path = tmp_path / "sales.xlsx"
        wb.save(file_path)

        parser = ExcelParser()
        result = await parser.parse(file_path)

        # Should extract content (as text and/or tables)
        assert len(result.chunks) > 0

        # Check that content contains the table data
        combined = " ".join(c.content for c in result.chunks)
        assert "Product" in combined


class TestExcelParserErrors:
    """Tests for error handling in Excel parser."""

    @pytest.mark.asyncio
    async def test_parse_nonexistent_file(
        self,
        mock_parser_settings: MagicMock,
    ) -> None:
        """Test handling of non-existent file."""
        from src.parsers.excel_parser import ExcelParser
        from src.parsers import CorruptFileError

        parser = ExcelParser()

        with pytest.raises(Exception):  # Could be CorruptFileError or FileNotFoundError
            await parser.parse(Path("/nonexistent/file.xlsx"))

    @pytest.mark.asyncio
    async def test_parse_corrupt_file(
        self,
        mock_parser_settings: MagicMock,
        tmp_path: Path,
    ) -> None:
        """Test handling of corrupt file."""
        from src.parsers.excel_parser import ExcelParser
        from src.parsers import CorruptFileError

        # Create a file with invalid content
        corrupt_file = tmp_path / "corrupt.xlsx"
        corrupt_file.write_bytes(b"This is not a valid Excel file")

        parser = ExcelParser()

        with pytest.raises(CorruptFileError):
            await parser.parse(corrupt_file)
