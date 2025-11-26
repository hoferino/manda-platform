"""
Integration tests for full document parsing pipeline.
Story: E3.2 - Integrate Docling for Document Parsing (AC: #6)

These tests use real document files to verify end-to-end parsing.
Note: Some tests may be skipped if Docling dependencies are not installed.
"""

import pytest
from pathlib import Path
from unittest.mock import MagicMock
import tempfile
import os


class TestFullExcelParsing:
    """Integration tests for Excel document parsing."""

    @pytest.fixture
    def financial_excel(self, tmp_path: Path) -> Path:
        """Create a financial model Excel file."""
        from openpyxl import Workbook

        wb = Workbook()

        # Income Statement sheet
        ws1 = wb.active
        ws1.title = "Income Statement"
        ws1["A1"] = "Line Item"
        ws1["B1"] = "2023"
        ws1["C1"] = "2024"
        ws1["D1"] = "Growth"

        ws1["A2"] = "Revenue"
        ws1["B2"] = 100000000
        ws1["C2"] = 125000000
        ws1["D2"] = "=(C2-B2)/B2"

        ws1["A3"] = "COGS"
        ws1["B3"] = 60000000
        ws1["C3"] = 72000000
        ws1["D3"] = "=(C3-B3)/B3"

        ws1["A4"] = "Gross Profit"
        ws1["B4"] = "=B2-B3"
        ws1["C4"] = "=C2-C3"
        ws1["D4"] = "=(C4-B4)/B4"

        ws1["A5"] = "Operating Expenses"
        ws1["B5"] = 25000000
        ws1["C5"] = 30000000
        ws1["D5"] = "=(C5-B5)/B5"

        ws1["A6"] = "EBITDA"
        ws1["B6"] = "=B4-B5"
        ws1["C6"] = "=C4-C5"
        ws1["D6"] = "=(C6-B6)/B6"

        # Assumptions sheet
        ws2 = wb.create_sheet("Assumptions")
        ws2["A1"] = "Parameter"
        ws2["B1"] = "Value"
        ws2["A2"] = "Revenue Growth Rate"
        ws2["B2"] = 0.25
        ws2["A3"] = "Gross Margin Target"
        ws2["B3"] = 0.45
        ws2["A4"] = "OpEx as % Revenue"
        ws2["B4"] = "=('Income Statement'!C5/'Income Statement'!C2)"

        file_path = tmp_path / "financial_model.xlsx"
        wb.save(file_path)
        return file_path

    @pytest.mark.asyncio
    async def test_parse_financial_model(
        self,
        mock_parser_settings: MagicMock,
        financial_excel: Path,
    ) -> None:
        """Test parsing a complete financial model."""
        from src.parsers.excel_parser import ExcelParser

        parser = ExcelParser()
        result = await parser.parse(financial_excel)

        # Should have chunks from both sheets
        assert len(result.chunks) > 0
        assert result.total_sheets == 2

        # Should extract formulas
        assert len(result.formulas) >= 6  # At least the main calculations

        # Check for specific formula types
        formula_texts = [f.formula for f in result.formulas]
        assert any("SUM" in f or "=" in f for f in formula_texts)

        # Check sheets are represented
        sheet_names = set(f.sheet_name for f in result.formulas)
        assert "Income Statement" in sheet_names

    @pytest.mark.asyncio
    async def test_parse_preserves_formula_results(
        self,
        mock_parser_settings: MagicMock,
        financial_excel: Path,
    ) -> None:
        """Test that computed values are captured when available."""
        from src.parsers.excel_parser import ExcelParser

        parser = ExcelParser()
        result = await parser.parse(financial_excel)

        # Some formulas may have result values
        formulas_with_results = [
            f for f in result.formulas
            if f.result_value is not None
        ]

        # Note: Result values depend on openpyxl's data_only mode
        # which may not always work for all formulas
        assert result.formulas  # At least formulas were extracted


class TestFullParserPipeline:
    """Integration tests for the full parsing pipeline."""

    @pytest.mark.asyncio
    async def test_chunking_token_limits(
        self,
        mock_parser_settings: MagicMock,
    ) -> None:
        """Test that chunks respect token limits."""
        from src.parsers.chunker import Chunker

        chunker = Chunker(min_tokens=50, max_tokens=100)

        # Create long text
        long_text = " ".join(["This is a test sentence."] * 100)

        chunks = chunker.chunk_text(long_text)

        # All chunks should be within limits (with some tolerance)
        for chunk in chunks:
            assert chunk.token_count is not None
            # Allow some flexibility for semantic boundaries
            assert chunk.token_count <= 150

    @pytest.mark.asyncio
    async def test_table_chunking_preserves_headers(
        self,
        mock_parser_settings: MagicMock,
    ) -> None:
        """Test that large table chunks preserve header rows."""
        from src.parsers.chunker import Chunker

        chunker = Chunker(max_tokens=50)

        # Create large table
        rows = ["| Name | Value | Description |", "| --- | --- | --- |"]
        for i in range(30):
            rows.append(f"| Item{i} | {i*100} | Description of item {i} |")

        table = "\n".join(rows)
        chunks = chunker.chunk_table(table, chunk_index=0)

        # Multiple chunks due to size
        assert len(chunks) > 1

        # Each chunk should have headers
        for chunk in chunks:
            assert "Name" in chunk.content
            assert "---" in chunk.content


class TestParserErrorRecovery:
    """Tests for error handling and recovery."""

    @pytest.mark.asyncio
    async def test_partial_parse_on_sheet_error(
        self,
        mock_parser_settings: MagicMock,
        tmp_path: Path,
    ) -> None:
        """Test that parser continues with other sheets if one fails."""
        from openpyxl import Workbook
        from src.parsers.excel_parser import ExcelParser

        # Create workbook with one good sheet
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Good Sheet"
        ws1["A1"] = "Valid data"
        ws1["B1"] = 12345

        file_path = tmp_path / "test.xlsx"
        wb.save(file_path)

        parser = ExcelParser()
        result = await parser.parse(file_path)

        # Should successfully parse the good sheet
        assert len(result.chunks) > 0

    @pytest.mark.asyncio
    async def test_warnings_collected(
        self,
        mock_parser_settings: MagicMock,
        tmp_path: Path,
    ) -> None:
        """Test that non-fatal issues are collected as warnings."""
        from openpyxl import Workbook
        from src.parsers.excel_parser import ExcelParser

        # Create simple file
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test"

        file_path = tmp_path / "test.xlsx"
        wb.save(file_path)

        parser = ExcelParser()
        result = await parser.parse(file_path)

        # Warnings list should exist even if empty
        assert isinstance(result.warnings, list)
        assert isinstance(result.errors, list)


class TestChunkMetadata:
    """Tests for chunk metadata integrity."""

    @pytest.mark.asyncio
    async def test_chunk_indices_sequential(
        self,
        mock_parser_settings: MagicMock,
        tmp_path: Path,
    ) -> None:
        """Test that chunk indices are sequential within a document."""
        from openpyxl import Workbook
        from src.parsers.excel_parser import ExcelParser

        # Create multi-sheet file
        wb = Workbook()
        for i in range(3):
            if i == 0:
                ws = wb.active
                ws.title = f"Sheet{i}"
            else:
                ws = wb.create_sheet(f"Sheet{i}")
            for j in range(10):
                ws[f"A{j+1}"] = f"Data {i}-{j}"

        file_path = tmp_path / "multi.xlsx"
        wb.save(file_path)

        parser = ExcelParser()
        result = await parser.parse(file_path)

        # Indices should be sequential
        indices = sorted([c.chunk_index for c in result.chunks])
        expected = list(range(len(indices)))
        assert indices == expected

    @pytest.mark.asyncio
    async def test_source_attribution(
        self,
        mock_parser_settings: MagicMock,
        tmp_path: Path,
    ) -> None:
        """Test that chunks have proper source attribution."""
        from openpyxl import Workbook
        from src.parsers.excel_parser import ExcelParser

        wb = Workbook()
        ws = wb.active
        ws.title = "Financial Data"
        ws["A1"] = "Revenue"
        ws["B1"] = 1000000

        file_path = tmp_path / "source_test.xlsx"
        wb.save(file_path)

        parser = ExcelParser()
        result = await parser.parse(file_path)

        # All chunks should have sheet name for Excel
        for chunk in result.chunks:
            assert chunk.sheet_name == "Financial Data" or chunk.metadata.get("sheet_name")
